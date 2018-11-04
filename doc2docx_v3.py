import os
import re
import openpyxl
import copy
import docx
from win32com import client as wc
#�����ȡ��������ݽṹ
PaparInfo = {'SNo':'','Sname':'','PaparTitle':'','Tname':'',\
            'Tlevel':'','Collega':'','Major':'','Class':'',\
             'Cnabstract':'','Cnkey':'','Enabstract':'',\
             'Enkey':'','Refercontext':'','Levle':'','PaparEnTitle':''
            }
#���ӱ��д������Ϣ��ѯ��
ExcleColumNamesTable = {'SNo':2,'Sname':3,'PaparTitle':4,'Tname':5,\
                        'Tlevel':6,'Collega':7,'Major':8,'Class':9,\
                        'Cnabstract':10,'Cnkey':11,'Enabstract':13,\
                        'Enkey':14,'Refercontext':15,'PaparEnTitle':12,'Levle':1
                        }
#��ȡ��ǰ�ļ��������е��ļ���·��
def getAllFolder(_pathName):
    try:
        folderlists = []
        for i,j,k in os.walk(_pathName):
            folderlists.append(i)
        return folderlists
    except  Exception as err:
        print('getAllFolder' + str(err))
#��ȡ��ǰ�ļ����µ��ļ�������
def getCurrentFilenames(folderdir):
    return os.listdir(folderdir)
#ɸ����DOC��docx�����ĵ�
def splitDocFromDocx(filenames):
    docFiles = []
    docxFiles = []
    for fn in filenames:
        if '.docx' in fn:
            docxFiles.append(fn)
        elif '.doc' in fn:
            docFiles.append(fn)
    return docFiles,docxFiles
'''
for i,j,k in myWalk:
    for ii in k:
        if ii != '':
            listFileNames.append(ii)
print(len(listFileNames))#output how many files in this flold!

for i in listFileNames:
    if '.docx' in i:
        listDocxFiles.append(i)
    elif '.doc' in i :
        listDocFiles.append(i)
print(listDocFiles)#output doc files'name
print(listDocxFiles)#output docx files's name
'''
#����Ƿ��ظ�ת������ȡ����ת���ļ����б�
def delRepeatFiles(docfiles,docxfiles):
    temp = []
    for i in docfiles:
        for j in docxfiles:
            if str(i) + 'x' == str(j):
                temp.append(i)
    for i in temp:
        docfiles.remove(i)
    return docfiles
#ɾ��~$��ʱ�ļ�
def delTempDocxFiles(docxfiles):
    temp = []
    for i in docxfiles:
        if '~$' in str(i):
            temp.append(i)
    for i in temp:
        docxfiles.remove(i)
    return docxfiles
#doc2docx,����ת���б�
def doc2docx(folderdir,docfiles,docxfiles):
    if docfiles == '':
        return docxfiles
    word = wc.Dispatch("Word.Application")
    for files in docfiles:
        try:
            temp_path = folderdir + '\\' + str(files)
            temp_doc = word.Documents.Open(temp_path)
            temp_doc.SaveAs(temp_path + 'x',16)
            docxfiles.append(str(files) + 'x')
            temp_doc.Close()
        except Exception as err:
            print('doc2docx: ' + files + str(err))
    word.Quit()
    return docxfiles
'''
for i in listDocFiles:#ѭ����DOC2DOCX
    temp_path = pathName + '\\' + str(i)
    temp_doc = word.Documents.Open(temp_path)
    temp_doc.SaveAs(temp_path + 'x',16)
    listDocxFiles.append(str(i) + 'x')
    temp_doc.Close()
word.Quit()
'''
#ɾ������Ӣ�Ŀո�
def cleanStringEnSpace(stringText):#��ϴ�ַ�����ǿո�
    return ''.join(stringText.split())

#ɾ������ȫ�ǿո�
def cleanStringFullSpace(stringText):#��ϴ�ַ���ȫ�ǿո�
    _stringText = ''
    for uchar in stringText:
        charcode = ord(uchar)
        if charcode == 12288 :
            continue
        _stringText += chr(charcode)
    return _stringText

def getText(filename):#��DOCXת��������Ȼ��Ϊ��λ���б�ṹ
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(cleanStringFullSpace(para.text.strip()))
    return fullText

#�������Ĺؼ��ʣ��������ո�ֽ�Ϊ�ֺ�
def changeSpace2Splite(stringText):
    _stringText = ''
    _stringTextList = stringText.split(' ')
    for i in _stringTextList:
        if i.strip() != '':
            _stringText = _stringText  + i + '��'
    return _stringText[0:len(_stringText)-1]

'''
tempText = []#��������
SNo = ''#ѧ��
Sname = ''#ѧ������
Collega = ''#ѧԺ
Major = ''#רҵ
Class = ''#רҵ
PaparTitle = ''#������Ŀ
Tname = ''#ָ����ʦ
Tlevel = ''#��ʦְ��
Cnabstract = ''#����ժҪ
Cnkey = ''#���Ĺؼ���
CnRefer = ''#���Ĳο�����
'''

#tempText = getText(r'D:\����\121304044.docx')
#��ȡ��Ŀ
def getPaparTitle(tempText):
    PaparTitle = ''
    for temp in tempText:
        if temp == '':
            continue
        temp = cleanStringEnSpace(temp)
        #temp = cleanStringFullSpace(temp)
        if '��Ŀ' in temp:
            PaparTitle = temp.replace('��Ŀ','')
            break
    if PaparTitle != '':
        if PaparTitle[0] == ':' or PaparTitle[0] =='��':
            PaparTitle = PaparTitle[1:]
        if PaparTitle[0] == '��' and PaparTitle[-1] == '��':
            PaparTitle = PaparTitle[1:-2]
    return PaparTitle
#��ȡѧԺ��רҵ
def getCollegaAndMajor(tempText):
    Collega = ''
    Major = ''
    Class = ''
    for temp in tempText:
        if temp == '':
            continue
        temp = cleanStringEnSpace(temp)
        if 'ѧԺ' in temp and 'רҵ' in temp and '��' in temp:
            Class = temp.split('��')[1].replace('\t','')
            _temp = temp.split('רҵ')[0].replace('\t','')
            _temp = _temp.split('ѧԺ')
            Collega = _temp[0]
            Major = _temp[1]
            break
    return Collega,Major,Class

#��ȡѧ��������ѧ��
def getSnameAndSNo(tempText):
    Sname = ''
    SNo = ''
    for temp in tempText:
        if temp == '':
            continue
        temp = cleanStringEnSpace(temp)
        if 'ѧ������' in temp :
            if 'ѧ��' in temp:
                Sname = temp.replace('ѧ������','')
                _temp = Sname.split('ѧ��')
                Sname = _temp[0].replace('\t','')
                SNo = _temp[1].replace('ѧ��','').replace('\t','')
                if Sname[0] == ':' or Sname[0] == '��':
                    Sname = Sname[1:]
                if SNo != '' and SNo[0] == ':' or SNo == '��':
                    SNo = SNo[1:]
                break
            else:
                Sname = temp.replace('ѧ������', '').replace(' ','')
                if Sname[0] == ':' or Sname[0] == '��':
                    Sname = Sname[1:]
                break
    if SNo == '':#���ѧ������һ�У������ٶ�ȡһ��
        for temp in tempText:
            if temp == '':
                continue
            temp = cleanStringEnSpace(temp)
            if 'ѧ��' in temp:
                SNo = temp.replace('ѧ��','').replace(' ','')
                if SNo[0] == ':' or SNo[0] == '��':
                    SNo = SNo[1:]
                break
    return Sname,SNo
#��ȡָ����ʦ
def getTname(tempText):
    Tname = ''
    for temp in tempText:
        if temp == '':
            continue
        temp = cleanStringEnSpace(temp)
        if 'ָ����ʦ' in temp:
            if 'ְ��' in temp:
                Tname = temp.replace('ָ����ʦ','')
                Tname = Tname.split('ְ��')[0].replace('\t','')
                break
            else:
                Tname = temp.replace('ָ����ʦ','')
                break
        elif 'ָ����ʦ' in temp:
            if 'ְ��' in temp:
                Tname = temp.replace('ָ����ʦ', '')
                Tname = Tname.split('ְ��')[0].replace('\t', '')
                break
            else:
                Tname = temp.replace('ָ����ʦ', '')
                break
    if Tname != '':
        if Tname[0] == ':' or Tname[0] == '��':
            Tname = Tname[1:]
    return Tname

#��ȡ��ʦְ��
def getTlevel(tempText):
    Tlevel = ''
    for temp in tempText:
        if temp == '':
            continue
        temp = cleanStringEnSpace(temp)
        if 'ְ��' in temp:
            Tlevel = temp[temp.find('ְ��')+2:]
            break
    if Tlevel != '':
        if Tlevel[0] == '��' or Tlevel[0] == ':':
            Tlevel = Tlevel[1:]
    return Tlevel

#��ȡ����ժҪ
def getCnabstract(tempText):
    Cnabstract = ''
    for i in range(len(tempText)):
        if tempText[i] == '':
            continue
        temp = cleanStringEnSpace(tempText[i])
        #temp = cleanStringFullSpace(temp)
        if '��ժҪ��' in temp:
            Cnabstract = temp.split('��ժҪ��')[1]
            j = i + 1
            while j < len(tempText):
                if tempText[j] == '':
                    break
                temp2 = cleanStringEnSpace(tempText[j])
                #temp2 = cleanStringFullSpace(temp2)
                if '���ؼ��ʡ�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��ʣ�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                elif '���ؼ��֡�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��֣�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                else:
                    Cnabstract = Cnabstract + '\n' + temp2
                    j = j + 1
                    continue
            break
        if '[ժҪ]' in temp:
            Cnabstract = temp.split('[ժҪ]')[1]
            j = i + 1
            while j < len(tempText):
                if tempText[j] == '':
                    break
                temp2 = cleanStringEnSpace(tempText[j])
                #temp2 = cleanStringFullSpace(temp2)
                if '���ؼ��ʡ�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��ʣ�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                elif '���ؼ��֡�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��֣�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                else:
                    Cnabstract = Cnabstract + '\n' + temp2
                    j = j + 1
                    continue
            break
        if 'ժҪ��' in temp:
            Cnabstract = temp.split('ժҪ��')[1]
            j = i + 1
            while j < len(tempText):
                if tempText[j] == '':
                    break
                temp2 = cleanStringEnSpace(tempText[j])
                # temp2 = cleanStringFullSpace(temp2)
                if '���ؼ��ʡ�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��ʣ�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                elif '���ؼ��֡�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��֣�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                else:
                    Cnabstract = Cnabstract + '\n' + temp2
                    j = j + 1
                    continue
            break
        if 'ժҪ:' in temp:
            Cnabstract = temp.split('ժҪ:')[1]
            j = i + 1
            while j < len(tempText):
                if tempText[j] == '':
                    break
                temp2 = cleanStringEnSpace(tempText[j])
                # temp2 = cleanStringFullSpace(temp2)
                if '���ؼ��ʡ�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��ʣ�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                elif '���ؼ��֡�' in temp2:
                    break
                elif '[�ؼ���]' in temp2:
                    break
                elif '�ؼ��֣�' in temp2:
                    break
                elif '�ؼ���:' in temp2:
                    break
                else:
                    Cnabstract = Cnabstract + '\n' + temp2
                    j = j + 1
                    continue
            break
    if Cnabstract != '':

        if Cnabstract[0] == ':' or Cnabstract[0] == '��':
            return Cnabstract[1:]
        else:
            return Cnabstract
    return Cnabstract

#��ȡ���Ĺؼ���
def getCnKey(tempText):
    Cnkey = ''
    i = 0
    for i in range(len(tempText)):
        temp = tempText[i].strip()
        if temp == '':
            continue

        if '���ؼ��ʡ�' in temp:
            temp = temp.split('���ؼ��ʡ�')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '[�ؼ���]' in temp:
            temp = temp.split('[�ؼ���]')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '�ؼ��ʣ�' in temp:
            temp = temp.split('�ؼ��ʣ�')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '�ؼ���:' in temp:
            temp = temp.split('�ؼ���:')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '���ؼ��֡�' in temp:
            temp = temp.split('���ؼ��֡�')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '[�ؼ���]' in temp:
            temp = temp.split('[�ؼ���]')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '�ؼ��֣�' in temp:
            temp = temp.split('�ؼ��֣�')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break
        elif '�ؼ���:' in temp:
            temp = temp.split('�ؼ���:')[1]
            temp2 = tempText[i + 1].strip()
            if temp2 != '' and '��һ��' not in temp2 and '����' not in temp2 and '����' not in temp2 and len(temp2)<10:
                temp = temp + tempText[i + 1]
            else:
                break

    if temp != '' and i != len(tempText):
        Cnkey = temp
        if ',' in Cnkey:
            Cnkey = cleanStringEnSpace(Cnkey.replace(',','��'))
        elif ';' in Cnkey:
            Cnkey = cleanStringEnSpace(Cnkey.replace(';','��'))
        elif ' ' in Cnkey:
            Cnkey = changeSpace2Splite(Cnkey)
        elif '��' in Cnkey:
            Cnkey = Cnkey.replace('��','��')
        elif '\\' in Cnkey:
            Cnkey  =  Cnkey.replace('\\','��')
    if Cnkey != '':
        if Cnkey[0] == ':' or Cnkey[0] == '��':
            return Cnkey[1:]
        else:
            return Cnkey
    return Cnkey

#��ȡ���Ĳο����׺�Ӣ����Ŀ
def getCnReferContextAndEnTitle(tempText,_enTitle):
    ReferContext = ''
    EnTitle = ''
    i = 0
    for i in range(len(tempText)-1,-1,-1):
        itemText = tempText[i].strip()
        if itemText == '':
            continue
        #itemText = cleanStringFullSpace(itemText)
        if '�ο�����' in itemText:
            i  = i + 1
            break
    ii = 0
    for ii in range(i,len(tempText),1):
        itemText = tempText[ii].strip()
        if itemText == '':
            continue
        #itemText == cleanStringFullSpace(itemText)
        if itemText[0] == '[' and str(itemText[1]).isdecimal() and \
                (str(itemText[2]).isdecimal()and itemText[3] == ']' or itemText[2] == ']'):
            ReferContext = itemText
            ii = ii + 1
            break
    iii = 0
    for iii in range(ii,len(tempText),1):
        itemText = tempText[iii].strip()
        #itemText == cleanStringFullSpace(itemText)
        if itemText == '':
            break
        elif itemText[0] == '[' and str(itemText[1]).isdecimal() and \
                (str(itemText[2]).isdecimal()and itemText[3] == ']' or itemText[2] == ']'):
            ReferContext = ReferContext + '��' + itemText
        else:
            ReferContext = ReferContext + itemText

    if _enTitle == '':
        iiii = 0#��ȡһ��Ӣ����Ŀ���˳�
        for iiii in range(iii,len(tempText),1):
            itemText = tempText[iiii].strip()
            if itemText == '':
                continue
            else:
                EnTitle = itemText
                break
    else:
        EnTitle = _enTitle
    return ReferContext,EnTitle

#��ȡӢ��ժҪ��Ӣ�Ĺؼ���
def getEnabstractAndKey(tempText):
    Enabstract = ''
    Enkey = ''
    try:
        i = 0
        for i in range(len(tempText)-1,-1,-1):#��λӢ��ժҪλ��
            itemText = tempText[i]
            if itemText.strip() == '':
                continue
            if '[Abstract]' in itemText:
                Enabstract = itemText.split('[Abstract]')[1]
                break
            elif '��Abstract��' in itemText:
                Enabstract = itemText.split('��Abstract��')[1]
                break
            elif '[abstract]' in itemText:
                Enabstract = itemText.split('[abstract]')[1]
                break
            elif '��abstract��' in itemText:
                Enabstract = itemText.split('��abstract��')[1]
                break
            elif 'abstract' in itemText:
                Enabstract = itemText.split('abstract')[1]
                break
            elif 'Abstract' in itemText:
                Enabstract = itemText.split('Abstract')[1]
                break

        if Enabstract != '':#��ȡժҪ��������
            ii = 0
            for ii in range(i+1,len(tempText),1):
                itemText = tempText[ii].strip()
                if itemText == '':
                    break
                if itemText[0] == '[' or itemText[0] == '��':
                    break
                elif 'key word' in itemText or 'key Word' in itemText or 'keyword' in itemText or 'keyWord' in itemText\
                        or 'Key word' in itemText or 'Key Word' in itemText or 'Keyword' in itemText or 'KeyWord' in itemText :
                    break
                else:
                    Enabstract = Enabstract + '\n' + itemText
            j = 0
            for j in range(ii,len(tempText),1):#��λӢ�Ĺؼ���λ��
                itemText = tempText[j].strip()
                if itemText == '':
                    continue
                if itemText[0] == '[':
                    Enkey = itemText.split(']')[1].strip()
                    break
                elif itemText[0] == '��':
                    Enkey = itemText.split('��')[1].strip()
                    break
                elif 'Key words' in itemText:
                    Enkey = itemText.replace('Key words','')
                    break
                elif 'Key Words' in itemText:
                    Enkey = itemText.replace('Key Words','')
                    break
                elif 'key words' in itemText:
                    Enkey = itemText.replace('key words','')
                    break
                elif 'key Words' in itemText:
                    Enkey = itemText.replace('key Words','')
                    break
                elif 'keywords' in itemText:
                    Enkey = itemText.replace('keywords','')
                    break
                elif 'Keywords' in itemText:
                    Enkey = itemText.replace('Keywords','')
                    break
                elif 'KeyWords' in itemText:
                    Enkey = itemText.replace('KeyWords','')
                    break
                elif 'keyWords' in itemText:
                    Enkey = itemText.replace('keyWords','')
                    break
                elif 'key word' in itemText:
                    Enkey = itemText.replace('key word','')
                    break
                elif 'key Word' in itemText:
                    Enkey = itemText.replace('key Word','')
                    break
                elif 'keyword' in itemText:
                    Enkey = itemText.replace('keyword','')
                    break
                elif 'keyWord' in itemText:
                    Enkey = itemText.replace('keyWord','')
                    break
                elif 'Key word' in itemText:
                    Enkey = itemText.replace('Key word','')
                    break
                elif 'Key Word' in itemText:
                    Enkey = itemText.replace('Key Word','')
                    break
                elif 'Keyword' in itemText:
                    Enkey = itemText.replace('Keyword','')
                    break
                elif 'KeyWord' in itemText:
                    Enkey = itemText.replace('KeyWord','')
                    break
            if Enkey != '':
                if j+1 < len(tempText) :#���һ�У����������ΪӢ�Ĺؼ���
                    itemText = tempText[j + 1].strip()
                    if itemText != '':
                        Enkey = Enkey + itemText
                if ';'in Enkey:
                    Enkey = '��'.join(Enkey.split(';'))
                elif ','in Enkey:
                    Enkey = '��'.join(Enkey.split(','))
                elif '��'in Enkey:
                    Enkey = '��'.join(Enkey.split('��'))
                elif '.'in Enkey:
                    Enkey = '��'.join(Enkey.split('.'))

        if Enabstract != '':#��ȥð��
            if Enabstract[0] == ':' or Enabstract[0] == '��':
                Enabstract = Enabstract[1:]
        if Enkey != '':
            if Enkey[0] == ':' or Enkey[0] =='��':
                Enkey = Enkey[1:]

    except Exception as err:
        print('getEnabstractAndKey' + filename + str(err))
    return Enabstract,Enkey

#�ж��Ƿ�������
def ischinese(strtemp):
    zhPattern = re.compile(u'[\u4e00-\u9fa5]+')
    match = zhPattern.search(strtemp)
    if match:
        return True
    else: return False

#��ȡӢר���ĵ���Ӣ����Ŀ
def getEnMajorTitle(tempText):
    PaparTitle = ''
    PaparEntitle = ''
    i = 0
    for i in range(len(tempText)):
        itemText = cleanStringEnSpace(tempText[i].strip())
        if itemText == '':
            continue
        if '��ҵ����' in itemText:
            for j in range(i+1,len(tempText),1):
                temp2 = tempText[j].strip()
                if temp2 == '':
                    continue
                else:
                    for k in range(j,len(tempText),1):
                        temp3 = tempText[k].strip()
                        if ischinese(temp3)== False and temp3 != '':
                            PaparEntitle = PaparEntitle + temp3
                        else:
                            break
                    for m in range(k,len(tempText),1):
                        temp4 = tempText[m].strip()
                        if temp4 == '':
                            continue
                        else:
                            PaparTitle = PaparTitle + temp4
                            if tempText[m + 1].strip() != '':
                                continue
                            else: break
                    break
            break
    return PaparTitle,PaparEntitle

'''
print(tempText)
print(getPaparTitle(tempText))
print(getSnameAndSNo(tempText))
print(getCollegaAndMajor(tempText))
print(getTname(tempText))
print(getTlevel(tempText))
print(getCnabstract(tempText))
print(getCnKey(tempText))
print(getCnReferContext(tempText))
print(getEnabstractAndKey(tempText))
'''
pathName = input('PLEASE INPUT THE PATH:\t')

listAllFolder = getAllFolder(pathName)
irow = 1
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()


for Cdir in listAllFolder:
    Cdirname = Cdir.split('\\')[-1]
    if 'һ������' in Cdirname:
        _levle = 'һ������'
    elif '��������' in Cdirname:
        _levle = '��������'
    else:
        _levle = 'һ������'
    listAllFolder = []
    listFileNames = []
    listDocFiles = []
    listDocxFiles = []
    listFileNames = getCurrentFilenames(Cdir)
    listFileNames = delTempDocxFiles(listFileNames)
    listDocFiles,listDocxFiles = splitDocFromDocx(listFileNames)
    if len(listDocFiles) == 0 and len(listDocxFiles) == 0:#û��DOC��DOCX�ļ�ֱ������
        continue
    print('���ڴ�����ȡ��Ϣ���ļ����ǣ�' + Cdir)
    listDocFiles = delRepeatFiles(listDocFiles,listDocxFiles)
    listDocxFiles = doc2docx(Cdir,listDocFiles,listDocxFiles)
    #print(len(listDocxFiles))
    for filename in listDocxFiles:
        try:
            tempText = getText(Cdir + '\\' + filename)
            #print(tempText)
            tempPaparInfo = copy.copy(PaparInfo)
            tempPaparInfo['Levle'] = _levle
            tempPaparInfo['PaparTitle'] = getPaparTitle(tempText)
            temp = getSnameAndSNo(tempText)
            tempPaparInfo['Sname'] = temp[0]
            tempPaparInfo['SNo'] = temp[1]
            if tempPaparInfo['SNo'] == '':
                tempPaparInfo['SNo'] = str(filename)
            temp = getCollegaAndMajor(tempText)
            tempPaparInfo['Collega'] = temp[0]
            tempPaparInfo['Major'] = temp[1]
            tempPaparInfo['Class'] = temp[2]
            tempPaparInfo['Tname'] = getTname(tempText)
            tempPaparInfo['Tlevel'] = getTlevel(tempText)
            tempPaparInfo['Cnabstract'] = getCnabstract(tempText)
            tempPaparInfo['Cnkey'] = getCnKey(tempText)
            temp = getCnReferContextAndEnTitle(tempText,tempPaparInfo['PaparEnTitle'])
            tempPaparInfo['Refercontext'] = temp[0]
            tempPaparInfo['PaparEnTitle'] = temp[1]
            temp = getEnabstractAndKey(tempText)
            tempPaparInfo['Enabstract'] = temp[0]
            tempPaparInfo['Enkey'] = temp[1]

            if tempPaparInfo['PaparTitle'] == '':
                temp = getEnMajorTitle(tempText)
                tempPaparInfo['PaparTitle'] = temp[0]
                tempPaparInfo['PaparEnTitle'] = temp[1]


            for ik,iv in tempPaparInfo.items():
                sheet.cell(row=irow, column=ExcleColumNamesTable.get(ik)).value = iv
            irow += 1
        #wb.save(Cdir + '\\' + Cdirname +'.xlsx' )
        except Exception as err:
            print(Cdir + filename + str(err))
            sheet.cell(row = irow,column = 2).value = str(filename)
            irow += 1
    print('             ' + Cdir + "������ϣ�")
    saveFilePath = pathName.split('\\')
    saveFileName = saveFilePath[-2] + saveFilePath[-1]
wb.save(pathName +'\\' + saveFileName + r'.xlsx')